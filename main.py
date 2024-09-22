import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime

file_name = 'Money Management Sheet.xlsx'

class money_manager:
    
    def __init__ (self, file_name):
        self.file_name = file_name
        self.load_workbook()
    
    def load_workbook(self):
        try:
            self.wb = openpyxl.load_workbook(self.file_name)
            self.sheet = self.wb.active
            print('>>>', self.file_name, "exists in the folder.")
        except FileNotFoundError:
            self.wb = Workbook()
            self.sheet = self.wb.active
            print('>>>', self.file_name, "does not exists in the folder.")
            print(">>> Creating File...")
            self.setup_headers()
            
    def setup_headers(self):
        headers = ['Date', 'Description', 'Category', 'Income', 'Expense', 'Balance']
        self.sheet.append(headers)
        self.wb.save(self.file_name)
        
    def log_transactions(self, description, category, income = 0, expense = 0):
        date = datetime.now().strftime('%Y - %m - %d')
        self.sheet.append([date,description, category, income, expense])
        self.wb.save(self.file_name)
        
    def calc_balance(self):
        total_income = sum(row[3] or 0 for row in self.sheet.iter_rows(min_row=2, values_only=True))
        total_expense = sum(row[4] or 0 for row in self.sheet.iter_rows(min_row=2, values_only=True))
        savings = total_income - total_expense
        print(">>>", savings, 'have been done till now.')

    def generate_report(self):
        try:
            wb: Workbook = openpyxl.load_workbook('Financial Report.xlsx')
            report_sheet = wb.active
            print('>>> Report.xlsx exists in the folder.')
        except FileNotFoundError:
            report_wb = Workbook()
            report_sheet = report_wb.active
            print(">>> Financial Report.xlsx does not exists in the folder.")
            print(">>> Creating File...")
            report_sheet.title = 'Financial Report'
        
        total_income = sum(row[3] or 0 for row in self.sheet.iter_rows(min_row=2, values_only=True))
        total_expense = sum(row[4] or 0 for row in self.sheet.iter_rows(min_row=2, values_only=True))
        savings = total_income - total_expense
        
        report_sheet.append([datetime.now().strftime('%Y - %m - %d')])
        report_sheet.append(['Total Income', total_income])
        report_sheet.append(['Total Expenses', total_expense])
        report_sheet.append(['Total savings', savings])
        
        report_wb.save('Financial Report.xlsx')
    
    def view_transaction(self):
        for row in self.sheet.iter_rows(min_row=1, values_only=True):
            print(row)
            print('---------------------------------')
            
    def delete_old_sheets(self):
        try:
            os.remove(self.file_name)
            print('Deleted',self.file_name, 'Sucessfully.')
        except FileNotFoundError as e:
            print(e,':',self.file_name, 'is removed already.')
        try:
            os.remove('Financial Report.xlsx')
            print('Deleted Financial Report.xlsx Sucessfully.')
        except FileNotFoundError as e:
            print(e,': Financial Report.xlsx is removed already.')

def useAgain():
    global ask
    while True:
            use = input('Do you wanna use the system again? (y/n)\n>>> ').lower()

            if not use.isalpha() or len(use) > 1 or use not in ['y','n']:
                print("Use 'y' for yes or 'n' for no.")
                continue
            elif use == 'y':
                break
            if use == 'n':
                print("Goodbye!")
                ask = False
                break
    

if __name__ == '__main__':
    manager = money_manager(file_name)
    
    ask = True
    
    while ask:
        print("""
What are you looking for?
1.Log Transaction
2.Calculate Balance
3.Generate Report
4.View Transaction History
5.Delete Old Transactions and Reports
""")
        ans = int(input(">>> "))
    
        if ans < 1 or ans > 5:
            continue
        elif ans == 1:
            description = (input('Description\n>>>'))
            category = (input('Category\n>>>'))
            income = int(input('Income\n>>>'))
            expenses = int(input('Expenses\n>>>'))
            
            manager.log_transactions(description, category, income, expenses)
            
        elif ans == 2:
            manager.calc_balance()
        
        elif ans == 3:
            manager.generate_report()
            
        elif ans == 4:
            manager.view_transaction()
            
        elif ans == 5:
            manager.delete_old_sheets()
        
        useAgain()